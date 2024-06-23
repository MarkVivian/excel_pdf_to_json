from datetime import datetime

import openpyxl


def get_cell_value(sheet, row, col):
    # Retrieve the value of the cell at the given row and column in the sheet
    cell_value = sheet.cell(row=row, column=col).value
    if isinstance(cell_value, datetime):
        try:
            cell_value = cell_value.strftime("%Y.%m.%d")
        except ValueError:
            pass

        # Check if date has "-" and validate first part
        if "/" in str(cell_value) or "\\" in str(cell_value):
            #     parts = str(cell_value).split(".")
            #     if len(parts[0]) != 4:
            #         print(f"Invalid date format (less than 4 characters before -): {cell_value}")
            # else:
            print(f"Date format missing '-': {cell_value}")

    return cell_value


def get_cell_fill_color(sheet, row, col):
    # Retrieve the background color of the cell at the given row and column in the sheet
    return sheet.cell(row=row, column=col).fill.start_color.index


class Excel:
    def __init__(self, file_path):
        # Initialize the Excel object with the file path of the Excel file
        self.file_path = file_path

    def process_excel(self):
        # Load the workbook from the given file path
        wb = openpyxl.load_workbook(self.file_path)

        # Select the active sheet in the workbook
        sheet = wb.active

        # Define the RGB hex code for the green color used in the background
        green_fill_color = 'FF00B050'  # ARGB format for green

        # Initialize the starting row
        row = 2
        # Initialize a list to hold all the processed data
        data = []

        # Loop through the rows starting from row 2 to the maximum row in the sheet
        # while row <= sheet.max_row:
        while row <= 8563:
            print(f"row number {row}")
            # Check if the cell in the 'name' column of the current row has a green background
            if get_cell_fill_color(sheet, row, 2) == green_fill_color:
                # Retrieve the 'all names', payroll and M+5 values from the current row
                main_data = {
                    "M+5": get_cell_value(sheet, row, 2) if get_cell_value(sheet, row, 2) is not None else "",
                    "first name": get_cell_value(sheet, row, 5) if get_cell_value(sheet, row, 5) is not None else "",
                    "middle name": get_cell_value(sheet, row, 6) if get_cell_value(sheet, row, 6) is not None else "",
                    "surname": get_cell_value(sheet, row, 7) if get_cell_value(sheet, row, 7) is not None else "",
                    "payroll number": get_cell_value(sheet, row, 4) if get_cell_value(sheet, row,
                                                                                      4) is not None else "",
                    "date of birth": get_cell_value(sheet, row, 9) if get_cell_value(sheet, row, 9) is not None else "",
                }
                # Initialize a list to hold the sub-row data related to the current main row
                sub_data = []

                # Move to the next row
                row += 1
                # Loop through the subsequent rows until another green-filled row is encountered
                while row <= sheet.max_row and get_cell_fill_color(sheet, row, 2) != green_fill_color:
                    # Retrieve the values from the current sub-row
                    sub_row_data = {
                        "M+5": get_cell_value(sheet, row, 2) if get_cell_value(sheet, row, 2) is not None else "",
                        "first name": get_cell_value(sheet, row, 5) if get_cell_value(sheet, row,
                                                                                      5) is not None else "",
                        "middle name": get_cell_value(sheet, row, 6) if get_cell_value(sheet, row,
                                                                                       6) is not None else "",
                        "surname": get_cell_value(sheet, row, 7) if get_cell_value(sheet, row, 7) is not None else "",
                        "gender": get_cell_value(sheet, row, 8) if get_cell_value(sheet, row, 8) is not None else "",
                        "date of birth": get_cell_value(sheet, row, 9) if get_cell_value(sheet, row,
                                                                                         9) is not None else "",
                        "relationship": get_cell_value(sheet, row, 10) if get_cell_value(sheet, row,
                                                                                         10) is not None else "",
                    }
                    # Append the sub-row data to the sub_data list
                    sub_data.append(sub_row_data)
                    # Move to the next row
                    row += 1

                # Append the main row data along with its related sub-rows to the data list
                data.append({
                    "employee": main_data,
                    "employee info": sub_data
                })
            else:
                # If the current row does not have a green background, move to the next row
                row += 1

        # sort data
        data = self.sort_data(data)

        # Return the processed data
        return data

    @staticmethod
    def sort_data(data):
        for item in data:
            employee = item["employee"]
            first_name = employee["first name"]
            payroll = employee["payroll number"]

            # check if first name is string
            if not isinstance(first_name, str):
                print(f"first name \n first name : {first_name} \n payroll : {payroll}")

        data_returned = sorted(data, key=lambda x: x["employee"].get("first name", " ").lower().strip(), reverse=False)
        return data_returned
